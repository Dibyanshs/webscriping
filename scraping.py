from bs4 import BeautifulSoup
import requests 
import openpyxl

excel=openpyxl.Workbook()
print(excel.sheetnames)
sheet=excel.active
sheet.title='top Rated Movies'
print(excel.sheetnames)
sheet.append(['Movie Rank','Movie name','year','IMDB ratng'])





try:
    source = requests.get('https://www.imdb.com/chart/toptv/') # url link of website
    source.raise_for_status()  # it find the error of the url

    soup= BeautifulSoup(source.text,'html.parser') # source.text=show the content of the website in text format 
   # print(soup)
    movies=soup.find('tbody',class_="lister-list").find_all('tr')
   # print(len(movies))    #It showes how many movies are there

    for movie in movies:    

        name=movie.find('td', class_="titleColumn").a.text  # It showes the name of movies

        rank=movie.find('td', class_="titleColumn").get_text(strip=True).split('.')[0]  # it show the rank

        year=movie.find('td', class_="titleColumn").span.text.strip('()')

        rating=movie.find('td',class_="ratingColumn imdbRating").strong.text

        
        print(rank, name, year, rating)
        sheet.append([rank, name, year, rating])
        
        


except Exception as e:
    print(e)
excel.save("Imdb movie rating.xlsx")

